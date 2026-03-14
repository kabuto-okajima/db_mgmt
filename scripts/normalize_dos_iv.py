#!/usr/bin/env python3
"""
Normalize DOS IV monthly Excel files into per-month CSVs and one combined CSV.

Output columns:
    year
    month
    basis
    fsc_or_place_of_birth
    visa_class
    issuances
    source_file

Notes:
  - Some monthly workbooks contain both FSC and POB sections.
  - Some monthly workbooks may contain duplicate keys even within the same basis.
    Those duplicates are aggregated by SUM, and the exact source rows/values are
    written to a log file.

Dependencies:
    pip install pandas openpyxl
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


RAW_DIR = Path("data/raw/dos_iv")
MONTHLY_OUT_DIR = Path("data/staging/dos_iv/monthly")
LOG_DIR = Path("data/staging/dos_iv/logs")
COMBINED_OUT_PATH = Path(
    "data/staging/dos_iv/dos_iv_fsc_or_place_of_birth_visa_class_monthly.csv"
)
AGGREGATION_LOG_PATH = LOG_DIR / "dos_iv_duplicate_aggregation.log"

MONTH_LOOKUP = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

KEY_COLS = [
    "year",
    "month",
    "basis",
    "fsc_or_place_of_birth",
    "visa_class",
]


@dataclass(frozen=True)
class ReportMeta:
    year: int
    month: int


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\xa0", " ")
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    text = re.sub(r"\s+", " ", text.strip())
    return text


def row_text(values: Tuple[Any, ...]) -> str:
    return normalize_text(" ".join(normalize_text(v) for v in values if v is not None))


def parse_issuances(value: Any) -> int:
    if value is None:
        raise ValueError("Issuances is null")
    if isinstance(value, (int, float)):
        return int(value)
    text = normalize_text(value).replace(",", "")
    if text == "":
        raise ValueError("Issuances is empty")
    return int(text)


def parse_report_metadata(ws, source_file: Path) -> ReportMeta:
    pattern = re.compile(r"([A-Za-z]+)\s+(\d{4})\s*\(\s*FY\s*(\d{4})\s*\)", re.I)

    for r in ws.iter_rows(
        min_row=1, max_row=25, min_col=1, max_col=12, values_only=True
    ):
        t = row_text(r)
        if not t:
            continue
        m = pattern.search(t)
        if m:
            month_name = m.group(1).lower()
            year = int(m.group(2))
            if month_name not in MONTH_LOOKUP:
                raise ValueError(f"Unknown month name in title: {m.group(1)}")
            return ReportMeta(year, MONTH_LOOKUP[month_name])

    fm = re.search(r"(\d{4})-(\d{2})", source_file.name)
    if fm:
        year = int(fm.group(1))
        month = int(fm.group(2))
        return ReportMeta(year, month)

    raise ValueError(
        f"Could not parse report metadata from worksheet or filename: {source_file}"
    )


def detect_basis(text: str) -> Optional[str]:
    t = text.lower()
    if "foreign state of chargeability" in t:
        return "FSC"
    if "place of birth" in t:
        return "POB"
    if re.search(r"\bfsc\b", t):
        return "FSC"
    if re.search(r"\bpob\b", t):
        return "POB"
    return None


def is_header_row_like(cells: Tuple[Any, Any, Any]) -> bool:
    c1 = normalize_text(cells[0]).lower()
    c2 = normalize_text(cells[1]).lower()
    c3 = normalize_text(cells[2]).lower()
    return (c2 == "visa class") and (c3 == "issuances") and (c1 != "")


def summarize_aggregation_group(group: pd.DataFrame) -> str:
    first = group.iloc[0]
    source_rows = ", ".join(str(int(v)) for v in group["source_row"].tolist())
    source_values = ", ".join(str(int(v)) for v in group["issuances"].tolist())
    total = int(group["issuances"].sum())

    return (
        f"source_file={first['source_file']} | "
        f"month={int(first['year'])}-{int(first['month']):02d} | "
        f"basis={first['basis']} | "
        f"fsc_or_place_of_birth={first['fsc_or_place_of_birth']} | "
        f"visa_class={first['visa_class']} | "
        f"source_rows=[{source_rows}] | "
        f"source_issuances=[{source_values}] | "
        f"summed_issuances={total}"
    )


def aggregate_duplicate_keys(
    df: pd.DataFrame,
    source_file: Path,
    aggregation_log_path: Path,
) -> pd.DataFrame:
    dupes = df.duplicated(subset=KEY_COLS, keep=False)
    if not dupes.any():
        return df

    duplicate_groups = (
        df.loc[dupes]
        .sort_values(KEY_COLS + ["source_row"])
        .groupby(KEY_COLS, dropna=False, sort=True)
    )

    aggregation_log_path.parent.mkdir(parents=True, exist_ok=True)
    with aggregation_log_path.open("a", encoding="utf-8") as logf:
        logf.write(f"\n=== Duplicate aggregation for {source_file.name} ===\n")
        for _, group in duplicate_groups:
            line = summarize_aggregation_group(group)
            print(f"[SUM DUPLICATES] {line}")
            logf.write(line + "\n")

    aggregated = (
        df.groupby(KEY_COLS, as_index=False, dropna=False)
        .agg(
            issuances=("issuances", "sum"),
            source_file=("source_file", "first"),
        )
        .sort_values(KEY_COLS)
        .reset_index(drop=True)
    )

    return aggregated


def parse_one_workbook(path: Path, aggregation_log_path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    meta = parse_report_metadata(ws, path)

    records: List[Dict[str, Any]] = []
    blank_streak = 0
    basis: Optional[str] = None

    for row_idx, r in enumerate(
        ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=3, values_only=True
        ),
        start=1,
    ):
        rtxt = row_text(r)

        maybe_basis = detect_basis(rtxt)
        if maybe_basis is not None:
            basis = maybe_basis

        if is_header_row_like(r):
            hb = detect_basis(rtxt)
            if hb is not None:
                basis = hb
            continue

        if rtxt == "":
            blank_streak += 1
            if blank_streak >= 30:
                break
            continue
        blank_streak = 0

        if basis is None:
            continue

        col1 = normalize_text(r[0])
        col2 = normalize_text(r[1])
        col3 = r[2]

        if not col1 or not col2:
            continue

        if col2.lower() == "visa class":
            continue

        try:
            issuances = parse_issuances(col3)
        except ValueError:
            continue

        records.append(
            {
                "year": meta.year,
                "month": meta.month,
                "basis": basis,
                "fsc_or_place_of_birth": col1,
                "visa_class": col2,
                "issuances": issuances,
                "source_file": path.name,
                "source_row": row_idx,
            }
        )

    if not records:
        raise ValueError(f"No records parsed from {path}")

    df = pd.DataFrame(records)
    df["basis"] = df["basis"].astype(str).str.strip()
    df["fsc_or_place_of_birth"] = df["fsc_or_place_of_birth"].astype(str).str.strip()
    df["visa_class"] = df["visa_class"].astype(str).str.strip()

    df = aggregate_duplicate_keys(
        df=df,
        source_file=path,
        aggregation_log_path=aggregation_log_path,
    )

    return df.sort_values(KEY_COLS).reset_index(drop=True)


def write_monthly_csv(df: pd.DataFrame, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    year = int(df["year"].iloc[0])
    month = int(df["month"].iloc[0])
    out_path = out_dir / f"{year}-{month:02d}_dos_iv_fsc_or_pob_visa_class.csv"
    df.to_csv(out_path, index=False)
    return out_path


def build_combined_csv(
    raw_dir: Path, monthly_out_dir: Path, combined_out_path: Path
) -> None:
    files = sorted(raw_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {raw_dir}")

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    AGGREGATION_LOG_PATH.write_text(
        "DOS IV duplicate aggregation log\n", encoding="utf-8"
    )

    all_frames: List[pd.DataFrame] = []

    for file_path in files:
        print(f"[READ] {file_path}")
        df = parse_one_workbook(file_path, AGGREGATION_LOG_PATH)
        monthly_csv_path = write_monthly_csv(df, monthly_out_dir)
        print(f"[WROTE MONTHLY CSV] {monthly_csv_path}")
        all_frames.append(df)

    combined = (
        pd.concat(all_frames, ignore_index=True)
        .sort_values(KEY_COLS)
        .reset_index(drop=True)
    )

    dupes = combined.duplicated(subset=KEY_COLS, keep=False)
    if dupes.any():
        duplicate_rows = combined.loc[dupes].sort_values(KEY_COLS)
        raise ValueError(
            "Duplicate keys found in combined dataset after monthly aggregation:\n"
            f"{duplicate_rows.to_string(index=False)}"
        )

    combined_out_path.parent.mkdir(parents=True, exist_ok=True)
    combined.to_csv(combined_out_path, index=False)

    months_count = combined[["year", "month"]].drop_duplicates().shape[0]
    print(f"[WROTE COMBINED CSV] {combined_out_path}")
    print(f"[ROWS] {len(combined):,}")
    print(f"[MONTHS] {months_count}")
    print(f"[WROTE LOG] {AGGREGATION_LOG_PATH}")


if __name__ == "__main__":
    build_combined_csv(
        raw_dir=RAW_DIR,
        monthly_out_dir=MONTHLY_OUT_DIR,
        combined_out_path=COMBINED_OUT_PATH,
    )
