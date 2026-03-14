#!/usr/bin/env python3
"""
Normalize DOS NIV monthly Excel files into per-month CSVs and one combined CSV.

Output columns:
    year
    month
    nationality
    visa_class
    issuances
    source_file

Dependencies:
    pip install pandas openpyxl
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import List, Dict, Any, Tuple

import pandas as pd
from openpyxl import load_workbook


RAW_DIR = Path("data/raw/dos_niv")
MONTHLY_OUT_DIR = Path("data/staging/dos_niv/monthly")
COMBINED_OUT_PATH = Path(
    "data/staging/dos_niv/dos_niv_nationality_visa_class_monthly.csv"
)

MONTH_LOOKUP = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

KEY_COLS = [
    "year",
    "month",
    "nationality",
    "visa_class",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text.strip())
    return text


def parse_report_metadata(ws, source_file: Path) -> Tuple[int, int]:
    """
    Example title:
      "Nonimmigrant Visa Issuances by Nationality July 2024 (FY 2024)"
    We only keep calendar year/month in the output.
    """
    pattern = re.compile(
        r"([A-Za-z]+)\s+(\d{4})\s+\(FY\s+(\d{4})\)",
        flags=re.IGNORECASE,
    )

    for row in ws.iter_rows(
        min_row=1, max_row=10, min_col=1, max_col=5, values_only=True
    ):
        joined = " ".join(normalize_text(cell) for cell in row if cell is not None)
        joined = normalize_text(joined)
        if not joined:
            continue

        match = pattern.search(joined)
        if match:
            month_name = match.group(1).lower()
            year = int(match.group(2))

            if month_name not in MONTH_LOOKUP:
                raise ValueError(f"Unknown month name in title: {match.group(1)}")

            month = MONTH_LOOKUP[month_name]
            return year, month

    filename_match = re.search(r"(\d{4})-(\d{2})", source_file.name)
    if filename_match:
        year = int(filename_match.group(1))
        month = int(filename_match.group(2))
        return year, month

    raise ValueError(
        f"Could not parse report metadata from worksheet or filename: {source_file}"
    )


def find_header_row(ws) -> int:
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        a = normalize_text(ws.cell(row=row_idx, column=1).value).lower()
        b = normalize_text(ws.cell(row=row_idx, column=2).value).lower()
        c = normalize_text(ws.cell(row=row_idx, column=3).value).lower()

        if a == "nationality" and b == "visa class" and c == "issuances":
            return row_idx

    raise ValueError(
        "Header row not found. Expected: Nationality | Visa Class | Issuances"
    )


def parse_issuances(value: Any) -> int:
    if value is None:
        raise ValueError("Issuances is null")

    if isinstance(value, (int, float)):
        return int(value)

    text = normalize_text(value).replace(",", "")
    if text == "":
        raise ValueError("Issuances is empty")

    return int(text)


def parse_one_workbook(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    year, month = parse_report_metadata(ws, path)
    header_row = find_header_row(ws)

    records: List[Dict[str, Any]] = []
    blank_streak = 0

    for row in ws.iter_rows(
        min_row=header_row + 1, min_col=1, max_col=3, values_only=True
    ):
        nationality = normalize_text(row[0])
        visa_class = normalize_text(row[1])
        issuances_raw = row[2]

        if not nationality and not visa_class and issuances_raw in (None, ""):
            blank_streak += 1
            if blank_streak >= 5:
                break
            continue

        blank_streak = 0

        if not nationality or not visa_class:
            continue

        try:
            issuances = parse_issuances(issuances_raw)
        except ValueError:
            continue

        records.append(
            {
                "year": year,
                "month": month,
                "nationality": nationality,
                "visa_class": visa_class,
                "issuances": issuances,
                "source_file": path.name,
            }
        )

    if not records:
        raise ValueError(f"No records parsed from {path}")

    df = pd.DataFrame(records)
    df["nationality"] = df["nationality"].str.strip()
    df["visa_class"] = df["visa_class"].str.strip()

    dupes = df.duplicated(subset=KEY_COLS, keep=False)
    if dupes.any():
        duplicate_rows = df.loc[dupes].sort_values(KEY_COLS)
        raise ValueError(
            f"Duplicate keys found in {path.name}:\n{duplicate_rows.to_string(index=False)}"
        )

    return df.sort_values(KEY_COLS).reset_index(drop=True)


def write_monthly_csv(df: pd.DataFrame, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)

    year = int(df["year"].iloc[0])
    month = int(df["month"].iloc[0])

    out_path = out_dir / f"{year}-{month:02d}_dos_niv_nationality_visa_class.csv"
    df.to_csv(out_path, index=False)
    return out_path


def build_combined_csv(
    raw_dir: Path, monthly_out_dir: Path, combined_out_path: Path
) -> None:
    files = sorted(raw_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {raw_dir}")

    all_frames: List[pd.DataFrame] = []

    for file_path in files:
        print(f"[READ] {file_path}")
        df = parse_one_workbook(file_path)

        monthly_csv_path = write_monthly_csv(df, monthly_out_dir)
        print(f"[WROTE MONTHLY CSV] {monthly_csv_path}")

        all_frames.append(df)

    combined = pd.concat(all_frames, ignore_index=True)
    combined = combined.sort_values(KEY_COLS).reset_index(drop=True)

    dupes = combined.duplicated(subset=KEY_COLS, keep=False)
    if dupes.any():
        duplicate_rows = combined.loc[dupes].sort_values(KEY_COLS)
        raise ValueError(
            "Duplicate keys found in combined dataset:\n"
            f"{duplicate_rows.to_string(index=False)}"
        )

    combined_out_path.parent.mkdir(parents=True, exist_ok=True)
    combined.to_csv(combined_out_path, index=False)

    print(f"[WROTE COMBINED CSV] {combined_out_path}")
    print(f"[ROWS] {len(combined):,}")
    print(f"[MONTHS] {combined[['year', 'month']].drop_duplicates().shape[0]}")


if __name__ == "__main__":
    build_combined_csv(
        raw_dir=RAW_DIR,
        monthly_out_dir=MONTHLY_OUT_DIR,
        combined_out_path=COMBINED_OUT_PATH,
    )
